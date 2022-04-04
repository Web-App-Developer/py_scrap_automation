import os
import sys
import openpyxl
from openpyxl import load_workbook

from selenium import webdriver
from bs4 import BeautifulSoup
import requests
import time
from selenium.webdriver.common.by import By

# 2017219944
# https://patentscope2.wipo.int/search/en/search.jsf

root_path = "https://patentscope2.wipo.int/search/en/search.jsf"


def getRequestDatas():
    file = "Patents_List1.xlsx"

    #load the work book
    wb_obj = load_workbook(filename = file)
    wsheet = wb_obj['20220326']
    dataDict = {}
    row_count = wsheet.max_row
    col_count = wsheet.max_column


    for row in range(2, row_count+1):
        key = wsheet.cell(row=row, column=3).value
        value = wsheet.cell(row=row, column=4).value
        dataDict[key] = value
    
    return dataDict


def scrap_one_page(dataDict):
    # foreach dataDict
    for key, value in dataDict.items():
        get_input_search_click_button(str(key))


def get_input_search_click_button(searchString):

    options = webdriver.ChromeOptions()
    options.add_argument("--window-size=1920,1080")
    options.binary_location = "C:\\Program Files\\Google\\Chrome\\Application\\chrome.exe"
    driver = webdriver.Chrome(options=options,executable_path=r'D:\\develop_experiences\\py_scraper_patentscope\\chromedriver.exe')
    driver.get(root_path)
    time.sleep(10)

    html = driver.execute_script("return document.body.innerHTML;")
    content = BeautifulSoup(html, "html.parser")
    input = driver.find_element(By.ID, "simpleSearchForm:fpSearch:input")
    slow_typing(input, searchString)
    time.sleep(0.1)
    
    # search button
    button = driver.find_element_by_id("simpleSearchForm:fpSearch:buttons").find_element_by_tag_name("button")
    button.click()
    time.sleep(0.2)

    # WO button
    driver.find_element_by_xpath("// span[contains(text(),\
    'WO')]").click()
    time.sleep(0.2)

    #tab
    driver.find_element_by_xpath("// a[contains(text(),\
    'Documents')]").click()
    time.sleep(10)

    #download
    driver.find_element_by_xpath("//td[contains(text(), \'Initial Publication with ISR')]/../td[4]/div/span/a").click()
    time.sleep(10) # changed the speed of network speed
    driver.quit()



def slow_typing(element, text):
    for character in text:
         element.send_keys(character)
         time.sleep(0.2)


def start():

    dataDict=getRequestDatas()
    print("dataDict",dataDict)
    scrap_one_page(dataDict)


start()
